from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter
wb=load_workbook('source_row.xlsx')
ws=wb.active
max_col=ws.max_column
for col in range(2,max_col+1):
    char=get_column_letter(col)
    # You can use ASCII table data to convert number to character. 
    # In ASCII table capital letters starts from position 65.
    # char=chr(65+col)
    # in this case you do not need import get_column_letter fuction from openpyxl.utils package
    hour=ws[char + str(2)].value
    rate=ws[char + str(3)].value
    if (type(hour)!=str and type(rate!=str)):
        salary=hour*rate
        ws[char+str(4)].value=salary
        print(round(salary,2))
wb.save('result.xlsx')
wb.close()