from openpyxl import Workbook, load_workbook 
wb=load_workbook('source.xlsx')
ws=wb.active
max_row=ws.max_row
for row in range(2,max_row+1):
    hour=ws['B' + str(row)].value
    rate=ws['C' + str(row)].value
    if (type(hour)!=str and type(rate!=str)):
        salary=hour*rate
        ws['D'+str(row)].value=salary
        print(round(salary,2))
wb.save('result.xlsx')
wb.close()