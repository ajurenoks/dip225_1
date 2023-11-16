from openpyxl import Workbook, load_workbook 
wb=load_workbook('saraksts1.xlsx')
ws=wb.active
max_row=ws.max_row
saraksts1 =[]
for row in range(1,max_row+1):
    rinda=[]
    A=ws['A' + str(row)].value
    

    saraksts1.append(A)
    

print("-------------------")   
wb=load_workbook('saraksts2.xlsx')
ws=wb.active
max_row=ws.max_row
saraksts2 =[]
for row in range(1,max_row+1):
    rinda=[]
    A=ws['A' + str(row)].value
   
    saraksts2.append(A)
    
result=[]
vs1=[]
for x in saraksts1:
    if x not in saraksts2:
        result.append(x)


for x in saraksts2:
    if x not in saraksts1:
        result.append(x)
        
print(result)